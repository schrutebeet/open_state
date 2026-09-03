import json
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO

import openpyxl

from civic_metrics.domain import DatasetPayload, ObservationCandidate, Period
from civic_metrics.genai_validation import _OUTPUT_SCHEMA, _candidate_dict, _payload_evidence


def test_json_payload_evidence_is_readable_and_reports_truncation() -> None:
    payload = DatasetPayload(
        dataset_code="sample",
        source_code="source",
        fetched_at=datetime.now(UTC),
        source_url="https://example.test/data.json",
        content_type="application/json",
        body=b'{"value": 123, "period": "2026-01"}',
        sha256="0" * 64,
    )

    evidence, truncated = _payload_evidence(payload, 10)

    assert evidence.startswith('{"value":')
    assert len(evidence) == 10
    assert truncated is True


def test_json_payload_evidence_uses_lossless_table_encoding_for_repeated_objects() -> None:
    payload = DatasetPayload(
        dataset_code="sample",
        source_code="source",
        fetched_at=datetime.now(UTC),
        source_url="https://example.test/data.json",
        content_type="application/json",
        body=(
            b'{"Data":[{"Fecha":1774994400000,"Anyo":2026,"Valor":10.0851},'
            b'{"Fecha":1767222000000,"Anyo":2026,"Valor":9.9755}]}'
        ),
        sha256="0" * 64,
    )

    evidence, truncated = _payload_evidence(payload, 1_000)
    compacted = json.loads(evidence)

    assert truncated is False
    assert compacted["Data"] == {
        "__civic_metrics_encoding__": "table",
        "columns": ["Fecha", "Anyo", "Valor"],
        "rows": [[1774994400000, 2026, 10.0851], [1767222000000, 2026, 9.9755]],
    }


def test_dataset_13_workbook_evidence_includes_only_tabla_1_5() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Tabla_1_5"
    sheet.append(["PERIODO", "SALDOS"])
    sheet.append(["202607", 10_000_000])
    sheet.append(["202607", 12_508_065.43])
    ignored = workbook.create_sheet("Other sheet")
    ignored.append(["must not be sent"])
    buffer = BytesIO()
    workbook.save(buffer)
    payload = DatasetPayload(
        dataset_code="sample",
        source_code="source",
        fetched_at=datetime.now(UTC),
        source_url="https://example.test/data.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        body=buffer.getvalue(),
        sha256="0" * 64,
    )

    evidence, truncated = _payload_evidence(payload, 1_000, dataset_id=13)

    assert truncated is False
    assert "sheet: Tabla_1_5" in evidence
    assert "SUM(SALDOS) grouped by PERIODO" in evidence
    assert "202607\t22508065.43" in evidence
    assert "Other sheet" not in evidence
    assert "must not be sent" not in evidence


def test_dataset_10_workbook_evidence_includes_only_tax_revenue_sheets() -> None:
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "Ingresos tributarios"
    source_row = [None] * 109
    source_row[0], source_row[1], source_row[2] = 2026, 6, "JUNIO"
    source_row[4], source_row[6] = -8_544_316, 13_934_443
    source_row[29], source_row[65], source_row[107] = 6_022_926, 444_050, 4_381_496
    source_row[108] = "must not be sent"
    first.append(source_row)
    ignored = workbook.create_sheet("Other sheet")
    ignored.append(["must not be sent"])
    buffer = BytesIO()
    workbook.save(buffer)
    payload = DatasetPayload(
        dataset_code="sample",
        source_code="source",
        fetched_at=datetime.now(UTC),
        source_url="https://example.test/data.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        body=buffer.getvalue(),
        sha256="0" * 64,
    )

    evidence, truncated = _payload_evidence(payload, 1_000, dataset_id=10)

    assert "sheet: Ingresos tributarios (structural projection)" in evidence
    assert "2026\t6\tJUNIO\t-8544316\t13934443\t6022926\t444050\t4381496" in evidence
    assert "Other sheet" not in evidence
    assert "must not be sent" not in evidence
    assert truncated is False


def test_candidate_is_serialized_without_losing_decimal_precision() -> None:
    candidate = ObservationCandidate(
        indicator_code="metric",
        source_code="source",
        dataset_code="sample",
        period=Period(
            start=date(2026, 1, 1),
            end=date(2026, 1, 31),
            label="2026-01",
            frequency="monthly",
        ),
        value=Decimal("123.4500"),
        unit="people",
    )

    assert _candidate_dict(candidate)["value"] == "123.4500"


def test_validation_schema_allows_only_the_requested_json_shape() -> None:
    schema = _OUTPUT_SCHEMA["schema"]

    assert schema["required"] == ["decision", "confidence", "description"]
    assert schema["properties"]["decision"]["enum"] == ["Valid", "Invalid"]
    assert schema["properties"]["confidence"] == {
        "type": "number",
        "minimum": 0,
        "maximum": 1,
    }
    assert schema["additionalProperties"] is False
