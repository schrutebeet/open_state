from __future__ import annotations

import re
from dataclasses import replace
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO

import openpyxl

from civic_metrics.catalog import DatasetDefinition, IndicatorDefinition
from civic_metrics.connectors.base import ConnectorContext
from civic_metrics.connectors.html_excel import HtmlExcelConnector
from civic_metrics.domain import DatasetPayload, ObservationCandidate
from civic_metrics.parsers.common import period_from_label


class SocialSecurityAffiliatesConnector(HtmlExcelConnector):
    """Extract monthly average affiliates from the official annual workbook."""

    connector_name = "social_security_affiliates"

    def fetch(self, dataset: DatasetDefinition, context: ConnectorContext) -> DatasetPayload:
        payload = super().fetch(dataset, context)
        return replace(
            payload,
            metadata={
                **payload.metadata,
                "history_periods": int(
                    dataset.config.get("history_periods", context.settings.max_history_periods)
                ),
            },
        )

    def extract(
        self,
        dataset: DatasetDefinition,
        payload: DatasetPayload,
        indicators: list[IndicatorDefinition],
    ) -> list[ObservationCandidate]:
        table_sheet = str(dataset.config.get("data_sheet", "Tabla_1_5"))
        book = openpyxl.load_workbook(BytesIO(payload.body), data_only=True, read_only=True)
        if table_sheet not in book.sheetnames:
            raise LookupError(f"Workbook does not contain expected sheet {table_sheet!r}")
        sheet = book[table_sheet]
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if headers is None:
            raise ValueError(f"Workbook sheet {table_sheet!r} has no header row")
        columns = {
            str(value).strip(): index for index, value in enumerate(headers) if value is not None
        }
        try:
            period_column = columns["PERIODO"]
            value_column = columns["SALDOS"]
        except KeyError as exc:
            raise ValueError(
                f"Workbook sheet {table_sheet!r} must contain PERIODO and SALDOS columns"
            ) from exc

        totals: dict[str, Decimal] = {}
        for row in rows:
            period_code = _period_code(row[period_column] if period_column < len(row) else None)
            value = row[value_column] if value_column < len(row) else None
            if period_code is None or value is None:
                continue
            totals[period_code] = totals.get(period_code, Decimal("0")) + Decimal(str(value))

        history_periods = int(payload.metadata.get("history_periods", 12))
        selected_periods = sorted(totals, reverse=True)[:history_periods]
        results: list[ObservationCandidate] = []
        for indicator in indicators:
            for period_code in selected_periods:
                period = period_from_label(
                    f"{period_code[:4]}-{period_code[4:]}",
                    indicator.frequency,
                )
                results.append(
                    ObservationCandidate(
                        indicator_code=indicator.code,
                        source_code=dataset.source,
                        dataset_code=dataset.code,
                        period=period,
                        value=totals[period_code].quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
                        unit=indicator.unit,
                        source_series=f"{table_sheet}!SUM(SALDOS) BY PERIODO",
                        source_url=payload.source_url,
                        metadata={
                            "data_sheet": table_sheet,
                            "period_code": period_code,
                            "aggregation": "SUM(SALDOS) grouped by PERIODO",
                        },
                    )
                )
        return results


def _period_code(value: object) -> str | None:
    text = str(value).strip()
    if re.fullmatch(r"20\d{4}", text):
        return text
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
        return text if re.fullmatch(r"20\d{4}", text) else None
    return None
