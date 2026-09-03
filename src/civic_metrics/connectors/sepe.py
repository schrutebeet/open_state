from __future__ import annotations

import calendar
import io
import os
import shutil
import subprocess
import tempfile
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

from civic_metrics.catalog import DatasetDefinition, IndicatorDefinition
from civic_metrics.connectors.html_excel import HtmlExcelConnector
from civic_metrics.domain import DatasetPayload, ObservationCandidate, Period
from civic_metrics.parsers.common import SPANISH_MONTHS, normalise_text, parse_decimal


class SepeRegisteredUnemploymentConnector(HtmlExcelConnector):
    connector_name = "sepe_registered_unemployment"

    @staticmethod
    def _is_missing(value: Any) -> bool:
        """Return True for None, empty strings, NaN and NaT."""

        if value is None:
            return True

        if isinstance(value, str):
            return not value.strip()

        try:
            missing = pd.isna(value)
        except (TypeError, ValueError):
            return False

        # pd.isna can theoretically return an array.
        return isinstance(missing, bool) and missing

    @staticmethod
    def _extract_month(values: list[Any]) -> int | None:
        """
        Extract the month using both the numeric month code and its name.

        Expected SEPE structure:
            column 1 -> month number
            column 2 -> month name
        """

        month_from_code: int | None = None
        month_from_name: int | None = None

        if values:
            raw_code = values[0]

            if not SepeRegisteredUnemploymentConnector._is_missing(raw_code):
                try:
                    month_from_code = int(float(raw_code))
                except (TypeError, ValueError):
                    month_from_code = None

                if month_from_code not in range(1, 13):
                    month_from_code = None

        if len(values) > 1:
            month_name = normalise_text(values[1])
            month_from_name = SPANISH_MONTHS.get(month_name)

        # Both should normally be identical.
        if (
            month_from_code is not None
            and month_from_name is not None
            and month_from_code != month_from_name
        ):
            raise ValueError(
                "SEPE month code and month name do not match: "
                f"code={month_from_code}, name={month_from_name}"
            )

        return month_from_code or month_from_name

    def extract(
        self,
        dataset: DatasetDefinition,
        payload: DatasetPayload,
        indicators: list[IndicatorDefinition],
    ) -> list[ObservationCandidate]:
        workbook = self._load_workbook(payload.body)
        sheet = workbook[workbook.sheetnames[0]]

        year_columns = self._year_columns(sheet)

        if not year_columns:
            raise LookupError("No year blocks found in SEPE workbook")

        today = date.today()

        # We try years from newest to oldest. This also handles a workbook
        # containing a future year block whose cells are still empty.
        selected: tuple[int, int, int, int, Decimal] | None = None

        for year in sorted(year_columns, reverse=True):
            value_column = year_columns[year]
            populated_months: list[tuple[int, int, Decimal]] = []

            for row_number, row in enumerate(
                sheet.iter_rows(values_only=True),
                start=1,
            ):
                values = list(row)
                month = self._extract_month(values)

                if month is None:
                    continue

                raw_value = (
                    values[value_column - 1]
                    if len(values) >= value_column
                    else None
                )

                if self._is_missing(raw_value):
                    continue

                try:
                    value = parse_decimal(raw_value)
                except (
                    InvalidOperation,
                    TypeError,
                    ValueError,
                ):
                    continue

                # Decimal("NaN") must never become an observation.
                if not value.is_finite():
                    continue

                # Defensive validation: do not publish a future month.
                if year > today.year:
                    continue

                if year == today.year and month > today.month:
                    continue

                populated_months.append(
                    (row_number, month, value)
                )

            if populated_months:
                # Select by month number, not by physical row order.
                row_number, month, value = max(
                    populated_months,
                    key=lambda item: item[1],
                )

                selected = (
                    year,
                    value_column,
                    row_number,
                    month,
                    value,
                )
                break

        if selected is None:
            raise LookupError(
                "No populated monthly unemployment value found "
                "in the SEPE workbook"
            )

        (
            latest_year,
            value_column,
            row_number,
            month,
            value,
        ) = selected

        period = Period(
            start=date(latest_year, month, 1),
            end=date(
                latest_year,
                month,
                calendar.monthrange(latest_year, month)[1],
            ),
            label=f"{latest_year}-{month:02d}",
            frequency="monthly",
        )

        results: list[ObservationCandidate] = []

        for indicator in indicators:
            if indicator.code != "registered_unemployment":
                continue

            results.append(
                ObservationCandidate(
                    indicator_code=indicator.code,
                    source_code=dataset.source,
                    dataset_code=dataset.code,
                    period=period,
                    value=value,
                    unit=indicator.unit,
                    source_series=(
                        f"{sheet.title}!"
                        f"R{row_number}C{value_column}"
                    ),
                    source_url=payload.source_url,
                    metadata={
                        "listing_url": payload.metadata.get(
                            "listing_url"
                        ),
                        "selected_link_text": payload.metadata.get(
                            "selected_link_text"
                        ),
                        "parser": (
                            "sepe_registered_unemployment_v2"
                        ),
                        "year": latest_year,
                        "month": month,
                        "excel_row": row_number,
                        "excel_column": value_column,
                    },
                )
            )

        return results

    @staticmethod
    def _year_columns(sheet: openpyxl.worksheet.worksheet.Worksheet) -> dict[int, int]:
        columns: dict[int, int] = {}
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 10)):
            for cell in row:
                text = normalise_text(cell.value)
                if not text.startswith("ano "):
                    continue
                try:
                    year = int(text.split()[-1])
                except ValueError:
                    continue
                # The year heading is placed over the registered-unemployment
                # column for that block in the official workbook.
                columns[year] = cell.column
        return columns


    @staticmethod
    def _find_libreoffice() -> str | None:
        configured = os.getenv("CIVIC_METRICS_LIBREOFFICE_PATH")
        candidates = [
            configured,
            shutil.which("libreoffice"),
            shutil.which("soffice"),
            os.path.join(
                os.getenv("PROGRAMFILES", r"C:\Program Files"),
                "LibreOffice",
                "program",
                "soffice.exe",
            ),
            os.path.join(
                os.getenv("PROGRAMFILES(X86)", r"C:\Program Files (x86)"),
                "LibreOffice",
                "program",
                "soffice.exe",
            ),
            "/Applications/LibreOffice.app/Contents/MacOS/soffice",
            "/usr/bin/libreoffice",
            "/usr/bin/soffice",
        ]
        for candidate in candidates:
            if candidate and Path(candidate).is_file():
                return str(candidate)
        return None

    @classmethod
    def _load_workbook(cls, body: bytes) -> openpyxl.Workbook:
        if body[:4] == b"PK\x03\x04":
            return openpyxl.load_workbook(BytesIO(body), data_only=True, read_only=True)

        else:
            try:
                file_stream = io.BytesIO(body)
                workbook = openpyxl.load_workbook(file_stream)
                return workbook
            except OSError:
                try:
                    df = pd.read_excel(io.BytesIO(body), engine="calamine", header=None, dtype=object)
                    df = df.astype(object).where(pd.notna(df), None)
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Sheet1"
                    for r in dataframe_to_rows(df, index=False, header=False):
                        ws.append(r)
                    return wb
                except Exception:
                    raise ValueError("Cannot read this Excel file.")
