from __future__ import annotations

import calendar
import re
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Iterable

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from civic_metrics.catalog import DatasetDefinition, IndicatorDefinition
from civic_metrics.connectors.direct_file import DirectFileConnector
from civic_metrics.connectors.html_excel import HtmlExcelConnector
from civic_metrics.domain import DatasetPayload, ObservationCandidate, Period
from civic_metrics.parsers.common import normalise_text, parse_decimal, period_from_label


class IgaeQuarterlyAccountsConnector(DirectFileConnector):
    """Parse IGAE's cumulative quarterly general-government accounts workbook."""

    connector_name = "igae_quarterly_accounts"

    _ROW_TERMS = {
        "general_government_revenue": "recursos no financieros",
        "general_government_expenditure": "empleos no financieros",
        "general_government_balance": "capacidad (+) o necesidad (-) de financiacion",
    }

    def extract(
        self,
        dataset: DatasetDefinition,
        payload: DatasetPayload,
        indicators: list[IndicatorDefinition],
    ) -> list[ObservationCandidate]:
        workbook = openpyxl.load_workbook(BytesIO(payload.body), data_only=True, read_only=True)
        if "Tabla1a" not in workbook.sheetnames:
            raise LookupError(f"Expected sheet 'Tabla1a'; found {workbook.sheetnames}")
        sheet = workbook["Tabla1a"]
        column, year, quarter = self._latest_populated_quarter(sheet)
        period = self._quarter_period(year, quarter)
        row_by_indicator = {
            code: self._find_row(sheet, term) for code, term in self._ROW_TERMS.items()
        }

        results: list[ObservationCandidate] = []
        for indicator in indicators:
            row = row_by_indicator.get(indicator.code)
            if row is None:
                continue
            value = parse_decimal(sheet.cell(row, column).value)
            results.append(
                ObservationCandidate(
                    indicator_code=indicator.code,
                    source_code=dataset.source,
                    dataset_code=dataset.code,
                    period=period,
                    value=value,
                    unit=indicator.unit,
                    source_series=f"Tabla1a!R{row}C{column}",
                    source_url=payload.source_url,
                    metadata={
                        "sheet": "Tabla1a",
                        "row": row,
                        "column": column,
                        "year": year,
                        "quarter": quarter,
                        "parser": "igae_quarterly_accounts_v1",
                        "source_unit": "million_eur",
                    },
                )
            )
        return results

    @classmethod
    def _latest_populated_quarter(cls, sheet: Worksheet) -> tuple[int, int, int]:
        current_year: int | None = None
        candidates: list[tuple[int, int, int]] = []
        key_rows = [cls._find_row(sheet, term) for term in cls._ROW_TERMS.values()]
        for column in range(3, sheet.max_column + 1):
            year_cell = sheet.cell(5, column).value
            if year_cell not in (None, ""):
                match = re.search(r"(?:19|20)\d{2}", str(year_cell))
                if match:
                    current_year = int(match.group(0))
            quarter_cell = normalise_text(sheet.cell(6, column).value)
            match = re.fullmatch(r"t([1-4])", quarter_cell)
            if current_year is None or match is None:
                continue
            quarter = int(match.group(1))
            if all(sheet.cell(row, column).value not in (None, "") for row in key_rows):
                candidates.append((current_year, quarter, column))
        if not candidates:
            raise LookupError("No populated IGAE quarterly column found")
        year, quarter, column = max(candidates, key=lambda item: (item[0], item[1]))
        return column, year, quarter

    @staticmethod
    def _find_row(sheet: Worksheet, term: str) -> int:
        target = normalise_text(term)
        for row in range(1, sheet.max_row + 1):
            label = " ".join(
                str(sheet.cell(row, column).value)
                for column in (1, 2)
                if sheet.cell(row, column).value not in (None, "")
            )
            if target in normalise_text(label):
                return row
        raise LookupError(f"Could not find IGAE row containing {term!r}")

    @staticmethod
    def _quarter_period(year: int, quarter: int) -> Period:
        start_month = 1 + (quarter - 1) * 3
        end_month = start_month + 2
        return Period(
            start=date(year, start_month, 1),
            end=date(year, end_month, calendar.monthrange(year, end_month)[1]),
            label=f"{year}-Q{quarter}",
            frequency="quarterly",
        )


class IgaeStateBudgetExecutionConnector(HtmlExcelConnector):
    """Parse the current-year State budget execution tables.

    The workbook is expressed in thousands of euros. Values are converted to
    millions of euros to match the public indicator catalog.
    """

    connector_name = "igae_state_budget_execution"
    _TO_MILLION = Decimal("0.001")

    def extract(
        self,
        dataset: DatasetDefinition,
        payload: DatasetPayload,
        indicators: list[IndicatorDefinition],
    ) -> list[ObservationCandidate]:
        workbook = openpyxl.load_workbook(BytesIO(payload.body), data_only=True, read_only=True)
        required = {"ING 002", "GTOS 001", "GTOS 004"}
        missing = required - set(workbook.sheetnames)
        if missing:
            raise LookupError(f"Missing expected IGAE sheets: {sorted(missing)}")

        income = workbook["ING 002"]
        expenditure = workbook["GTOS 001"]
        chapters = workbook["GTOS 004"]
        period = self._period_from_title(income)
        year = period.end.year

        income_total = self._find_row(income, exact_compact="totales")
        expenditure_total = self._find_row(expenditure, exact_compact="totales")
        interest_row = self._find_row(chapters, contains="gastos financieros")

        values: dict[str, tuple[Decimal, str]] = {
            "state_budget_revenue_forecast": self._cell_by_header(
                income, income_total, year, "previsiones presupuestarias"
            ),
            "state_budget_revenue_recognised": self._cell_by_header(
                income, income_total, year, "derechos reconocidos netos"
            ),
            "state_budget_final_appropriation": self._cell_by_header(
                expenditure, expenditure_total, year, "creditos definitivos"
            ),
            "state_budget_expenditure_recognised": self._cell_by_header(
                expenditure, expenditure_total, year, "obligaciones reconocidas netas"
            ),
            "state_interest_expenditure": self._cell_by_header(
                chapters, interest_row, year, "obligaciones reconocidas netas"
            ),
        }

        results: list[ObservationCandidate] = []
        for indicator in indicators:
            matched = values.get(indicator.code)
            if matched is None:
                continue
            raw_value, source_series = matched
            results.append(
                ObservationCandidate(
                    indicator_code=indicator.code,
                    source_code=dataset.source,
                    dataset_code=dataset.code,
                    period=period,
                    value=raw_value * self._TO_MILLION,
                    unit=indicator.unit,
                    source_series=source_series,
                    source_url=payload.source_url,
                    metadata={
                        "listing_url": payload.metadata.get("listing_url"),
                        "selected_link_text": payload.metadata.get("selected_link_text"),
                        "parser": "igae_state_budget_execution_v1",
                        "source_unit": "thousand_eur",
                        "conversion": "value / 1000",
                    },
                )
            )
        return results

    @staticmethod
    def _period_from_title(sheet: Worksheet) -> Period:
        title = " ".join(
            str(sheet.cell(row, column).value)
            for row in range(1, min(sheet.max_row, 5) + 1)
            for column in range(1, min(sheet.max_column, 8) + 1)
            if sheet.cell(row, column).value not in (None, "")
        )
        return period_from_label(title, "monthly")

    @staticmethod
    def _compact(value: object) -> str:
        return re.sub(r"[^a-z0-9]", "", normalise_text(value))

    @classmethod
    def _find_row(
        cls,
        sheet: Worksheet,
        *,
        exact_compact: str | None = None,
        contains: str | None = None,
    ) -> int:
        contains_normalised = normalise_text(contains) if contains else None
        for row in range(1, sheet.max_row + 1):
            value = sheet.cell(row, 1).value
            if exact_compact and cls._compact(value) == exact_compact:
                return row
            if contains_normalised and contains_normalised in normalise_text(value):
                return row
        raise LookupError(
            f"Could not find row in {sheet.title}: exact={exact_compact!r} contains={contains!r}"
        )

    @classmethod
    def _cell_by_header(
        cls,
        sheet: Worksheet,
        row: int,
        year: int,
        header_term: str,
    ) -> tuple[Decimal, str]:
        year_by_column = cls._forward_filled_years(sheet, header_rows=range(1, 4))
        target = normalise_text(header_term)
        matches: list[int] = []
        for column in range(2, sheet.max_column + 1):
            header = " ".join(
                str(sheet.cell(header_row, column).value)
                for header_row in range(1, 4)
                if sheet.cell(header_row, column).value not in (None, "")
            )
            if year_by_column.get(column) == year and target in normalise_text(header):
                matches.append(column)
        if not matches:
            raise LookupError(
                f"No {year} column containing {header_term!r} in sheet {sheet.title}"
            )
        column = min(matches)
        value = parse_decimal(sheet.cell(row, column).value)
        return value, f"{sheet.title}!R{row}C{column}"

    @staticmethod
    def _forward_filled_years(
        sheet: Worksheet,
        *,
        header_rows: Iterable[int],
    ) -> dict[int, int]:
        years: dict[int, int] = {}
        current: int | None = None
        for column in range(1, sheet.max_column + 1):
            for row in header_rows:
                value = sheet.cell(row, column).value
                if value in (None, ""):
                    continue
                match = re.fullmatch(r"(?:19|20)\d{2}", str(value).strip())
                if match:
                    current = int(match.group(0))
                    break
            if current is not None:
                years[column] = current
        return years
