from __future__ import annotations

import calendar
import re
from datetime import date
from decimal import Decimal
from io import BytesIO

import openpyxl

from civic_metrics.catalog import DatasetDefinition, IndicatorDefinition
from civic_metrics.connectors.html_excel import HtmlExcelConnector
from civic_metrics.domain import DatasetPayload, ObservationCandidate, Period
from civic_metrics.parsers.common import normalise_text, parse_decimal


class SocialSecurityMinimumSupplementsConnector(HtmlExcelConnector):
    """Extract the national total number of pensions with a minimum supplement."""

    connector_name = "social_security_minimum_supplements"

    def extract(
        self,
        dataset: DatasetDefinition,
        payload: DatasetPayload,
        indicators: list[IndicatorDefinition],
    ) -> list[ObservationCandidate]:
        workbook = openpyxl.load_workbook(BytesIO(payload.body), data_only=True, read_only=True)
        sheet_name = str(dataset.config.get("data_sheet", "Min_número_%"))
        if sheet_name not in workbook.sheetnames:
            raise LookupError(f"Workbook does not contain expected sheet {sheet_name!r}")
        value, row_number = self._national_total(workbook[sheet_name])
        period = self._period_from_payload(payload)
        return [
            ObservationCandidate(
                indicator_code=indicator.code,
                source_code=dataset.source,
                dataset_code=dataset.code,
                period=period,
                value=value,
                unit=indicator.unit,
                source_series=f"{sheet_name}!R{row_number}C2",
                source_url=payload.source_url,
                metadata={
                    "sheet": sheet_name,
                    "section": "AMBOS SEXOS",
                    "row": "Total",
                    "column": "TOTAL PENSIONES / Número",
                    "listing_url": payload.metadata.get("listing_url"),
                    "selected_link_text": payload.metadata.get("selected_link_text"),
                },
            )
            for indicator in indicators
        ]

    @staticmethod
    def _national_total(sheet: openpyxl.worksheet.worksheet.Worksheet) -> tuple[Decimal, int]:
        in_both_sexes = False
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = list(row)
            label = normalise_text(values[0] if values else "")
            if label.startswith("ambos sexos"):
                in_both_sexes = True
                continue
            if in_both_sexes and label == "total":
                value = values[1] if len(values) > 1 else None
                if value is None:
                    raise ValueError("Total pensions number is blank")
                return parse_decimal(value), row_number
        raise LookupError("Could not find AMBOS SEXOS / Total / TOTAL PENSIONES number")

    @staticmethod
    def _period_from_payload(payload: DatasetPayload) -> Period:
        combined = " ".join(
            [payload.source_url, str(payload.metadata.get("selected_link_text", ""))]
        )
        match = re.search(r"MIN(20\d{2})(0[1-9]|1[0-2])", combined, re.IGNORECASE)
        if not match:
            raise ValueError(f"Could not infer minimum-supplement period from {combined!r}")
        year, month = int(match.group(1)), int(match.group(2))
        return Period(
            start=date(year, month, 1),
            end=date(year, month, calendar.monthrange(year, month)[1]),
            label=f"{year}-{month:02d}",
            frequency="monthly",
        )
