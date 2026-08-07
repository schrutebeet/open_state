from __future__ import annotations

import calendar
import re
from datetime import date
from decimal import Decimal
from io import BytesIO

import openpyxl

from civic_metrics.catalog import DatasetDefinition, IndicatorDefinition
from civic_metrics.connectors.html_excel import HtmlExcelConnector
from civic_metrics.domain import DatasetPayload, ObservationCandidate, Period
from civic_metrics.parsers.common import normalise_text, parse_decimal


class SocialSecurityPensionsConnector(HtmlExcelConnector):
    """Dedicated parser for the INSS monthly pension workbooks.

    These books use hierarchical headers and time-series layouts that cannot be
    mapped safely with the generic Excel label matcher.
    """

    connector_name = "social_security_pensions"

    def extract(
        self,
        dataset: DatasetDefinition,
        payload: DatasetPayload,
        indicators: list[IndicatorDefinition],
    ) -> list[ObservationCandidate]:
        workbook = openpyxl.load_workbook(BytesIO(payload.body), data_only=True, read_only=True)
        period = self._period_from_payload(payload)
        handlers = {
            "social_security_pension_series": self._extract_series,
            "social_security_pension_payroll": self._extract_payroll,
            "social_security_pensioners": self._extract_pensioners,
        }
        handler = handlers.get(dataset.code)
        if handler is None:
            raise ValueError(f"Unsupported pension workbook dataset {dataset.code}")
        values = handler(workbook)
        return self._build_candidates(dataset, payload, indicators, period, values)

    @staticmethod
    def _period_from_payload(payload: DatasetPayload) -> Period:
        combined = " ".join(
            [
                payload.source_url,
                str(payload.metadata.get("selected_link_text", "")),
            ]
        )
        match = re.search(r"(?:PTAS|ICONCEPTOS|S)(20\d{2})(0[1-9]|1[0-2])", combined, re.I)
        if not match:
            raise ValueError(f"Could not infer pension workbook period from {combined!r}")
        year, month = int(match.group(1)), int(match.group(2))
        return Period(
            start=date(year, month, 1),
            end=date(year, month, calendar.monthrange(year, month)[1]),
            label=f"{year}-{month:02d}",
            frequency="monthly",
        )

    @staticmethod
    def _extract_series(workbook: openpyxl.Workbook) -> dict[str, tuple[Decimal, str]]:
        sheet = workbook["S_Total"]
        latest_row: tuple[int, list[object]] | None = None
        current_year: int | None = None
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = list(row)
            # The workbook contains a second percentage-change table below the
            # absolute series. Stop when that section begins.
            if values and normalise_text(values[0]) == "periodo" and row_number > 10:
                break
            if values and isinstance(values[0], (int, float)):
                current_year = int(values[0])
            month_label = values[1] if len(values) > 1 else None
            total_count = values[2] if len(values) > 2 else None
            if current_year is None or not isinstance(month_label, str) or total_count in (None, ""):
                continue
            latest_row = (row_number, values)
        if latest_row is None:
            raise LookupError("No populated monthly row found in S_Total")
        row_number, values = latest_row
        return {
            "pension_count": (parse_decimal(values[2]), f"S_Total!R{row_number}C3"),
            "average_pension": (parse_decimal(values[4]), f"S_Total!R{row_number}C5"),
            "average_retirement_pension": (
                parse_decimal(values[10]),
                f"S_Total!R{row_number}C11",
            ),
        }

    @staticmethod
    def _extract_payroll(workbook: openpyxl.Workbook) -> dict[str, tuple[Decimal, str]]:
        sheet = workbook["Importe"]
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = list(row)
            if values and normalise_text(values[0]) == "total sistema":
                # Under the first header group, "Total pensiones", the first
                # numeric column is the total payroll in millions of euros.
                return {
                    "pension_monthly_payroll": (
                        parse_decimal(values[1]),
                        f"Importe!R{row_number}C2",
                    )
                }
        raise LookupError("Could not find Total sistema in pension payroll workbook")

    @staticmethod
    def _extract_pensioners(workbook: openpyxl.Workbook) -> dict[str, tuple[Decimal, str]]:
        sheet = workbook["Resumen de datos"]
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = list(row)
            if values and normalise_text(values[0]) == "numero de pensionistas":
                return {
                    "pensioner_count": (
                        parse_decimal(values[1]),
                        f"Resumen de datos!R{row_number}C2",
                    )
                }
        raise LookupError("Could not find Número de pensionistas in workbook")

    @staticmethod
    def _build_candidates(
        dataset: DatasetDefinition,
        payload: DatasetPayload,
        indicators: list[IndicatorDefinition],
        period: Period,
        values: dict[str, tuple[Decimal, str]],
    ) -> list[ObservationCandidate]:
        results: list[ObservationCandidate] = []
        for indicator in indicators:
            matched = values.get(indicator.code)
            if matched is None:
                continue
            value, source_series = matched
            results.append(
                ObservationCandidate(
                    indicator_code=indicator.code,
                    source_code=dataset.source,
                    dataset_code=dataset.code,
                    period=period,
                    value=value,
                    unit=indicator.unit,
                    source_series=source_series,
                    source_url=payload.source_url,
                    metadata={
                        "listing_url": payload.metadata.get("listing_url"),
                        "selected_link_text": payload.metadata.get("selected_link_text"),
                        "parser": "social_security_pensions_v1",
                    },
                )
            )
        return results
